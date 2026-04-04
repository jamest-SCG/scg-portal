#!/usr/bin/env python3
"""
import_submissions.py — SCG PM Portal Helper Script

Takes the portal's export CSV and pastes submission data into the master
Excel workbook's JOB_DETAIL sheet (columns W through AH).

Usage:
    python import_submissions.py <export_csv> <master_excel>

Example:
    python import_submissions.py SCG_PM_Submissions_2026-02-26.csv "SCG Master Workbook.xlsx"

Column mapping (JOB_DETAIL sheet):
    Column A  = Job No.
    Column W  = Feb-26 Billings
    Column X  = Mar-26 Billings
    Column Y  = Apr-26 Billings
    Column Z  = May-26 Billings
    Column AA = Jun-26 Billings
    Column AB = Jul-26 Billings
    Column AC = Aug-26 Billings
    Column AD = Sep-26 Billings
    Column AE = Oct-26 Billings
    Column AF = Nov-26 Billings
    Column AG = Dec-26 Billings
    Column AH = PM Est Cost-to-Cmplt Override ($)

Requirements:
    pip install openpyxl
"""

import sys
import csv
import os

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


# CSV column order (matches portal export)
CSV_FIELDS = [
    'job_no', 'feb_26', 'mar_26', 'apr_26', 'may_26', 'jun_26', 'jul_26',
    'aug_26', 'sep_26', 'oct_26', 'nov_26', 'dec_26', 'ctc_override'
]

# Excel column letters for billing months + CTC override (W through AH)
EXCEL_COLS = ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH']

# The sheet name in the master workbook
SHEET_NAME = 'JOB_DETAIL'


def col_letter_to_index(letter):
    """Convert Excel column letter(s) to 1-based index."""
    result = 0
    for char in letter:
        result = result * 26 + (ord(char.upper()) - ord('A') + 1)
    return result


def main():
    if len(sys.argv) != 3:
        print("Usage: python import_submissions.py <export_csv> <master_excel>")
        print("Example: python import_submissions.py SCG_PM_Submissions_2026-02-26.csv workbook.xlsx")
        sys.exit(1)

    csv_path = sys.argv[1]
    excel_path = sys.argv[2]

    if not os.path.exists(csv_path):
        print(f"Error: CSV file not found: {csv_path}")
        sys.exit(1)

    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found: {excel_path}")
        sys.exit(1)

    # Read CSV export
    submissions = {}
    with open(csv_path, 'r', newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            job_no = row.get('job_no', '').strip()
            if not job_no:
                continue
            values = []
            for field in CSV_FIELDS[1:]:  # Skip job_no
                val = row.get(field, '').strip()
                if val == '' or val == '0':
                    values.append(None)
                else:
                    try:
                        values.append(float(val))
                    except ValueError:
                        values.append(None)
            submissions[job_no] = values

    print(f"Read {len(submissions)} job submissions from CSV.")

    # Open Excel workbook
    wb = load_workbook(excel_path)

    if SHEET_NAME not in wb.sheetnames:
        # Try case-insensitive match
        sheet_match = None
        for name in wb.sheetnames:
            if name.upper() == SHEET_NAME.upper():
                sheet_match = name
                break
        if sheet_match:
            ws = wb[sheet_match]
        else:
            print(f"Error: Sheet '{SHEET_NAME}' not found in workbook.")
            print(f"Available sheets: {', '.join(wb.sheetnames)}")
            sys.exit(1)
    else:
        ws = wb[SHEET_NAME]

    # Find the Job No. column (usually column A)
    job_no_col = None
    header_row = 1

    # Search first 10 rows for the header
    for row_idx in range(1, 11):
        for col_idx in range(1, 20):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val and 'job' in str(cell_val).lower() and 'no' in str(cell_val).lower():
                job_no_col = col_idx
                header_row = row_idx
                break
        if job_no_col:
            break

    if not job_no_col:
        print("Warning: Could not find 'Job No.' header. Assuming column A, row 1.")
        job_no_col = 1
        header_row = 1

    print(f"Found Job No. column at column {job_no_col}, header row {header_row}")

    # Build map of job_no -> row
    job_rows = {}
    for row_idx in range(header_row + 1, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=job_no_col).value
        if cell_val:
            job_rows[str(cell_val).strip()] = row_idx

    # Write submission data
    updated = 0
    not_found = []

    col_indices = [col_letter_to_index(c) for c in EXCEL_COLS]

    for job_no, values in submissions.items():
        if job_no in job_rows:
            row_idx = job_rows[job_no]
            for i, val in enumerate(values):
                if i < len(col_indices):
                    cell = ws.cell(row=row_idx, column=col_indices[i])
                    if val is not None:
                        cell.value = val
                    else:
                        cell.value = None
            updated += 1
        else:
            not_found.append(job_no)

    # Save
    output_path = excel_path
    wb.save(output_path)
    print(f"\nResults:")
    print(f"  Updated: {updated} jobs")
    if not_found:
        print(f"  Not found in workbook ({len(not_found)}): {', '.join(not_found)}")
    print(f"  Saved to: {output_path}")
    print("\nDone!")


if __name__ == '__main__':
    main()
